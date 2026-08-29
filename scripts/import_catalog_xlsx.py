#!/usr/bin/env python
"""Importa la hoja Catálogo maestro a professionalCatalog.json.

Uso:
  uv run --with openpyxl python scripts/import_catalog_xlsx.py archivo.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CATALOG_SOURCE = ROOT / "src" / "catalog" / "equipmentCatalog.js"
OUTPUT = ROOT / "src" / "catalog" / "professionalCatalog.json"
SHEET_NAME = "Catálogo maestro"
REQUIRED_HEADERS = (
    "ID interno",
    "Fabricante",
    "Modelo",
    "Código ODDO",
    "Peso (kg)",
    "Consumo nominal (W)",
    "Consumo máximo (W)",
    "Carga térmica (W)",
    "URL oficial",
    "Fuente / ficha técnica",
    "Fecha de revisión",
    "Estado",
    "Notas",
)
VALID_STATES = {"pendiente", "revisado", "validado", "descatalogado"}


def expected_ids() -> list[str]:
    source = CATALOG_SOURCE.read_text(encoding="utf-8")
    ids = re.findall(r"\bid:\s*'([^']+)'", source)
    if len(ids) != 32 or len(set(ids)) != len(ids):
        raise ValueError("No se pudieron identificar exactamente los 32 IDs oficiales del catálogo técnico.")
    return ids


def clean_text(value, *, nullable=False):
    if value is None:
        return None if nullable else ""
    text = str(value).strip()
    return text or (None if nullable else "")


def nullable_number(value, label, product_id):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        raise ValueError(f"{product_id}: {label} debe ser un número o quedar vacío.")
    try:
        number = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{product_id}: {label} debe ser un número o quedar vacío.") from error
    if number < 0:
        raise ValueError(f"{product_id}: {label} no puede ser negativo.")
    return int(number) if number.is_integer() else number


def nullable_date(value, product_id):
    if value in (None, ""):
        return None
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = str(value).strip()
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as error:
        raise ValueError(f"{product_id}: Fecha de revisión debe usar AAAA-MM-DD o quedar vacía.") from error


def nullable_url(value, product_id):
    text = clean_text(value, nullable=True)
    if text is None:
        return None
    parsed = urlparse(text)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError(f"{product_id}: URL oficial no válida: {text}")
    return text


def read_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Falta la hoja '{SHEET_NAME}'.")
    sheet = workbook[SHEET_NAME]
    headers = [clean_text(cell.value) for cell in sheet[1]]
    if "SKU fabricante" in headers:
        raise ValueError("El Excel todavía contiene la columna eliminada 'SKU fabricante'.")
    missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing_headers:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing_headers)}")
    positions = {header: headers.index(header) for header in REQUIRED_HEADERS}

    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        product_id = clean_text(values[positions["ID interno"]])
        if not product_id:
            continue
        state = clean_text(values[positions["Estado"]]).lower()
        if state not in VALID_STATES:
            raise ValueError(f"{product_id}: Estado no válido: {state or '(vacío)'}. Key: {', '.join(sorted(VALID_STATES))}")
        oddo = clean_text(values[positions["Código ODDO"]])
        if not oddo:
            raise ValueError(f"{product_id}: Código ODDO debe contener XXXXX o el código interno real.")
        rows.append({
            "id": product_id,
            "fabricante": clean_text(values[positions["Fabricante"]]),
            "modelo": clean_text(values[positions["Modelo"]]),
            "codigoODDO": oddo,
            "pesoKg": nullable_number(values[positions["Peso (kg)"]], "Peso", product_id),
            "consumoNominalW": nullable_number(values[positions["Consumo nominal (W)"]], "Consumo nominal", product_id),
            "consumoMaximoW": nullable_number(values[positions["Consumo máximo (W)"]], "Consumo máximo", product_id),
            "cargaTermicaW": nullable_number(values[positions["Carga térmica (W)"]], "Carga térmica", product_id),
            "urlOficial": nullable_url(values[positions["URL oficial"]], product_id),
            "fuente": clean_text(values[positions["Fuente / ficha técnica"]], nullable=True),
            "fechaRevision": nullable_date(values[positions["Fecha de revisión"]], product_id),
            "estadoValidacion": state,
            "notas": clean_text(values[positions["Notas"]], nullable=True),
        })
    return rows


def validate_ids(rows: list[dict], official_ids: list[str]):
    imported_ids = [row["id"] for row in rows]
    duplicates = sorted({item for item in imported_ids if imported_ids.count(item) > 1})
    missing = [item for item in official_ids if item not in imported_ids]
    unknown = [item for item in imported_ids if item not in official_ids]
    if duplicates or missing or unknown:
        raise ValueError(
            "Los IDs del Excel no coinciden con el catálogo oficial. "
            f"Duplicados: {duplicates or 'ninguno'}. "
            f"Faltan: {missing or 'ninguno'}. "
            f"Desconocidos: {unknown or 'ninguno'}."
        )


def main():
    parser = argparse.ArgumentParser(description="Importa el catálogo profesional desde Excel.")
    parser.add_argument("xlsx", type=Path, help="Ruta al Excel maestro")
    args = parser.parse_args()
    if not args.xlsx.is_file():
        raise ValueError(f"No existe el Excel: {args.xlsx}")

    official_ids = expected_ids()
    rows = read_rows(args.xlsx)
    validate_ids(rows, official_ids)
    by_id = {row["id"]: row for row in rows}
    ordered_rows = [by_id[product_id] for product_id in official_ids]
    OUTPUT.write_text(json.dumps(ordered_rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "ok": True,
        "productos": len(ordered_rows),
        "codigos_oddo_pendientes": sum(row["codigoODDO"] == "XXXXX" for row in ordered_rows),
        "salida": str(OUTPUT),
    }, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise SystemExit(1)
